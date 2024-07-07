import time
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException
import pandas as pd
from openpyxl import load_workbook

# 로그인 정보
username = 'hkn230717b'
password = 'hkn23071&'
login_url = 'https://www.mobon.net/main/m2/'
data_url = 'https://manage.mobon.net/report/daily_list'  # 실제 데이터 보고서 URL

# 요일 한국어 변환
def get_korean_day_of_week(date):
    days = ['월요일', '화요일', '수요일', '목요일', '금요일', '토요일', '일요일']
    return days[date.weekday()]

# 웹 드라이버 설정 (webdriver-manager 사용)
service = ChromeService(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)

def login_and_navigate_to_data_page(driver, username, password):
    try:
        driver.get(login_url)
        #time.sleep(5)  # 페이지가 로드되는 시간을 늘림

        # 아이디와 비밀번호 입력
        driver.find_element(By.ID, 'member_id').send_keys(username)
        driver.find_element(By.ID, 'member_passwd').send_keys(password)
        
        # JavaScript를 사용하여 로그인 버튼 클릭
        login_button = driver.find_element(By.ID, 'login')
        driver.execute_script("arguments[0].click();", login_button)
        
        time.sleep(5)  # 로그인 후 페이지 로드 대기

        # 로그인 성공 여부 확인
        if "로그인 실패" in driver.page_source:
            print("로그인 실패")
            return False
        else:
            print("로그인 성공")

            # 로그인 후 새로운 탭을 열어서 데이터 페이지로 이동
            driver.execute_script("window.open('');")
            driver.switch_to.window(driver.window_handles[1])
            driver.get(data_url)
            time.sleep(10)  # 데이터 페이지 로드 대기

            # "SKIP" 버튼 찾기
            try:
                skip_button = driver.find_element(By.CSS_SELECTOR, "a.btn-skip[data-marktype='coachMarkWrap_report_hkn230717b']")
                skip_button.click()
                print("SKIP 버튼 클릭 완료")
            except NoSuchElementException:
                print("SKIP 버튼이 없음")

            print("데이터 페이지로 이동 완료. 현재 URL:", driver.current_url)
            return True
    except Exception as e:
        print(f"로그인 또는 데이터 페이지 이동 중 오류 발생: {e}")
        return False

def get_data_for_dates(driver, dates):
    try:
        # 명시적 대기 추가
        wait = WebDriverWait(driver, 2)

        # 테이블 데이터 추출
        rows = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#tblExport tbody tr")))
        data_list = []

        for date in dates:
            for row in rows:
                row_date = row.find_element(By.CSS_SELECTOR, "td:nth-child(1)").text
                if row_date == date.strftime('%m-%d'):
                    view_count = row.find_element(By.CSS_SELECTOR, "td:nth-child(2)").text
                    click_count = row.find_element(By.CSS_SELECTOR, "td:nth-child(3)").text
                    ad_spent = row.find_element(By.CSS_SELECTOR, "td:nth-child(5)").text
                    conversions = row.find_element(By.CSS_SELECTOR, "td:nth-child(8)").text
                    total_sales = row.find_element(By.CSS_SELECTOR, "td:nth-child(16)").text
                    totalSalesRoas = row.find_element(By.CSS_SELECTOR, "td:nth-child(17)").text
                    day_of_week = get_korean_day_of_week(date)
                    data = {
                        '날짜': f"{row_date} ({day_of_week})",
                        '노출수': view_count,
                        '클릭수': click_count,
                        '소진금액': ad_spent,
                        '전환수': conversions,
                        '총매출': total_sales,
                        '총ROAS': totalSalesRoas
                    }
                    data_list.append(data)
                    break

        if not data_list:
            print("지정된 날짜의 데이터를 찾을 수 없습니다.")
            return None
        else:
            # 날짜 순으로 정렬
            data_list.sort(key=lambda x: datetime.strptime(x['날짜'].split()[0], '%m-%d'))
            return data_list

    except Exception as e:
        print(f"데이터 추출 중 오류 발생: {e}")
        return None

def save_to_excel(data_list, filename='data.xlsx'):
    try:
        # 기존 Excel 파일 열기
        wb = load_workbook(filename=filename)
        sheet = wb.active
        
        # 데이터를 Excel에 추가
        for r_idx, data in enumerate(data_list, start=2):
            sheet.cell(row=r_idx, column=1, value=data['날짜'])
            sheet.cell(row=r_idx, column=2, value=data['노출수'])
            sheet.cell(row=r_idx, column=3, value=data['클릭수'])
            sheet.cell(row=r_idx, column=4, value=data['소진금액'])
            sheet.cell(row=r_idx, column=5, value=data['전환수'])
            sheet.cell(row=r_idx, column=6, value=data['총매출'])
            sheet.cell(row=r_idx, column=7, value=data['총ROAS'])
        
        # Excel 파일 저장
        wb.save(filename)
        print(f"데이터가 {filename} 파일에 추가되었습니다.")
    except Exception as e:
        print(f"Excel 파일 저장 중 오류 발생: {e}")

try:
    if login_and_navigate_to_data_page(driver, username, password):
        today = datetime.now()
        if today.weekday() == 0:  # 현재 요일이 월요일인 경우
            dates = [today - timedelta(days=i) for i in range(1, 4)]  # 금, 토, 일 날짜
        else:
            dates = [today - timedelta(1)]  # 어제 날짜

        data_list = get_data_for_dates(driver, dates)
        if data_list:
            save_to_excel(data_list, filename='data.xlsx')
finally:
    driver.quit()
